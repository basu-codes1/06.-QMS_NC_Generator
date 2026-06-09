import time
import openpyxl as op
from openpyxl.styles import Border, Side, Alignment
from pathlib import PurePosixPath, PureWindowsPath, Path
from colorama import Fore, Style, init
import shutil
import os
import xlwings as xw
import pyperclip3
import sys
init(autoreset=True)
thin_side = Side(border_style="thin", color="000000")
border_style = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
left_alignment = Alignment(horizontal='left', vertical='center')

TOOL_NAME = """+--------------------------+
| NC Report Generator V1P0 |
+--------------------------+
"""

SIGN_OFF = """+-------------------------------------------------------------+
| Tool By: Basavaraja NT(3328) & DurgaPrasad(11468) ~Team QA. |
+-------------------------------------------------------------+"""

sys.stdout.write(f"{Fore.LIGHTCYAN_EX}{TOOL_NAME}{Style.RESET_ALL}")
file1_path = input("Please provide the input file path: ").strip()
file1_path = file1_path.replace('"', '')
file1_path = str(PurePosixPath(PureWindowsPath(file1_path)))
template_file_name = "T1-B-FM-QA-0305-02 Nonconformance report.xlsx"
summary_sheet = "NC Summary.xlsx"
while True:
    try:
        if file1_path.lower() == "exit":
            sys.stdout.write(f"{Fore.LIGHTMAGENTA_EX}{SIGN_OFF}{Style.RESET_ALL} \n")
            time.sleep(0.3)
            sys.exit()
        if file1_path.endswith(".xlsx"):
            if os.path.exists(file1_path):
                wb1 = op.load_workbook(file1_path)  #1st workbook
                sheets = wb1.sheetnames
                if len(sheets) == 2:
                    sh1 = wb1["Sheet1"]
                    max_rw = sh1.max_row
                    nc_count = 0
                    nc_list = []
                    row_number = []  # DP
                    for i in range(10, max_rw + 1):
                        g_value = sh1.cell(i, 7).value
                        if g_value is not None:
                            if g_value.strip() == "NC":
                                row_number.append(i)  # DP
                                nc_count += 1

                    if nc_count == 0:
                        sys.stdout.write(f"{Fore.LIGHTGREEN_EX}No NC found.{Style.RESET_ALL}\n")
                    else:
                        print(f"Total {nc_count} NC found.")
                        for nc in range(nc_count):
                            nc_list.append(f"NC_{nc + 1}")
                        nc_list.pop(0)

                        cwd = os.getcwd()
                        cwd = str(PurePosixPath(PureWindowsPath(cwd)))
                        template_path = cwd + "/Template/" + template_file_name
                        template_path = str(PurePosixPath(PureWindowsPath(template_path)))

                        if os.path.exists(template_path):
                            shutil.copy(template_path, cwd)
                            xl_path = cwd + f"/{template_file_name}"  #2nd file

                            standard = sh1.cell(4, 3).value
                            location = sh1.cell(4, 8).value

                            iqa_num = sh1.cell(5, 3).value
                            auditee = sh1.cell(5, 8).value

                            department = sh1.cell(6, 3).value
                            auditor = sh1.cell(6, 8).value

                            date = sh1.cell(7, 3).value .strftime("%d-%b-%Y")
                            observations = sh1.cell(row_number[0], 2).value  # DP

                            wb2 = op.load_workbook(xl_path)
                            wb2["Sheet1"].title = "NC_1"
                            if len(nc_list) > 0:
                                for i in nc_list:
                                    wb2.create_sheet(i)
                            # print(nc_list)
                            sh2 = wb2["NC_1"]
                            sh2.cell(5, 2).value = standard
                            sh2.cell(5, 4).value = location
                            sh2.cell(6, 2).value = iqa_num
                            sh2.cell(6, 4).value = "NC1"   #Need to add this cell in for loop later
                            sh2.cell(7, 2).value = department
                            sh2.cell(7, 4).value = auditee
                            sh2.cell(8, 2).value = date
                            sh2.cell(8, 4).value = auditor
                            sh2.cell(11, 2).value = observations


                            # sh2.cell(11, 2).value = observations

                            sh2.cell(12, 2).value = auditor
                            sh2.cell(12, 5).value = date

                            sh2.cell(16, 2).value = auditee
                            # sh2.cell(16, 5).value = date

                            sh2.cell(20, 2).value = auditee
                            # sh2.cell(20, 5).value = date

                            sh2.cell(25, 2).value = auditee
                            # sh2.cell(25, 5).value = date

                            sh2.cell(30, 3).value = auditor
                            # sh2.cell(30, 5).value = date

                            # sh2.cell(35, 5).value = date
                            # sh2.cell(37, 5).value = f"Date: {date}"
                            wb2.save(xl_path)
                            print(f"{Fore.BLUE}Extraction of NC_1 completed.{Style.RESET_ALL}")

                            # ---COPYING SHEET - DP Start
                            # Run Excel in background (no window)
                            if len(nc_list) > 0:
                                app = xw.App(visible=False)
                                app.screen_updating = False
                                app.calculation = 'manual'
                                wb = app.books.open(xl_path)

                                # Source sheet
                                source_sheet = wb.sheets["NC_1"]
                                # source_sheet.range("B11").value = observations
                                # source_sheet.range("D6").value = "NC_1"
                                # print(f"{Fore.BLUE}Extraction of NC_1 completed.{Style.RESET_ALL}")
                                # Loop through your existing sheets
                                for sheet_name in nc_list:
                                    target_sheet = wb.sheets[sheet_name]
                                    # Copy everything from source to target
                                    source_sheet.api.Cells.Copy(target_sheet.api.Cells)
                                    observations_nc = sh1.cell(row_number[nc_list.index(sheet_name) + 1], 2).value
                                    target_sheet.range("B11").value = observations_nc
                                    target_sheet.range("D6").value = sheet_name
                                    print(f"{Fore.BLUE}Extraction of {sheet_name} completed.{Style.RESET_ALL}")
                                wb.save(xl_path)
                                app.calculation = 'automatic'
                                app.screen_updating = True
                                app.api.CutCopyMode = False
                                wb.close()
                                app.quit()
                                pyperclip3.clear()
                                pyperclip3.copy(file1_path)
                            # # ---COPYING SHEET - DP END HERE
                            file_save_name = template_file_name.split(".")[0]
                            file_save_name = file_save_name + f"_{department}" + ".xlsx"
                            if os.path.exists(file_save_name):
                                os.remove(file_save_name)
                            os.rename(xl_path, file_save_name)
                            dir_name = os.path.dirname(xl_path)
                            file_save_name = dir_name + "/" + file_save_name

                            sys.stdout.write(f"{Fore.BLUE}Updating summary sheet...{Style.RESET_ALL}\n")
                            summary = cwd + f"/{summary_sheet}"

                            if os.path.exists(summary):
                                wb_summary = op.load_workbook(summary)
                                sh_summary = wb_summary["NC Summary"]
                                nc_added = False
                                for nc_range in range(10, max_rw +1 ):
                                    if sh1.cell(nc_range, 7).value is not None:
                                        if sh1.cell(nc_range, 7).value.strip() == "NC" and sh1.cell(nc_range, 2).value is not None:
                                            nc_obs = sh1.cell(nc_range, 2).value.strip() + "_" + department  #sh1 observation
                                            summary_mx_row = sh_summary.max_row
                                            found = 0
                                            for sum_row in range(3, summary_mx_row + 1):
                                                if sh_summary.cell(sum_row, 6).value is not None:
                                                    sum_obs = sh_summary.cell(sum_row, 6).value.strip() + "_" + sh_summary.cell(sum_row, 5).value
                                                    if nc_obs == sum_obs:
                                                        found = 1
                                                        break
                                            if found == 0:
                                                nc_added = True
                                                c_blank = 0
                                                for blank in range(2, 5000):
                                                    if sh_summary.cell(blank, 3).value is None:
                                                        c_blank = blank
                                                        break
                                                sh_summary.cell(c_blank, 2).value = "NC"
                                                sh_summary.cell(c_blank, 3).value = standard
                                                sh_summary.cell(c_blank, 4).value = location
                                                sh_summary.cell(c_blank, 5).value = department
                                                sh_summary.cell(c_blank, 6).value = sh1.cell(nc_range, 2).value
                                if nc_added:
                                    for row in sh_summary.iter_rows(min_row=3, max_row=500, min_col=1, max_col=1):
                                        for cell in row:
                                            cell.value = None

                                    c_blank = 0
                                    for blank in range(2, 5000):
                                        if sh_summary.cell(blank, 3).value is None:
                                            c_blank = blank
                                            break
                                    for sl_num in range(3, c_blank):
                                        sh_summary.cell(sl_num, 1).value = sl_num - 2
                                mx_rw = sh_summary.max_row
                                for row in sh_summary.iter_rows(min_row=2, max_row= 30, min_col=1, max_col=15):
                                    for cell in row:
                                        cell.border = border_style
                                        cell.alignment = left_alignment
                                wb_summary.save(summary)
                                sys.stdout.write(
                                    f"{Fore.LIGHTGREEN_EX}NC Report generated and Saved in: {Style.RESET_ALL}{file_save_name}\n")
                                sys.stdout.write(
                                    f"{Fore.LIGHTGREEN_EX}Summary workbook Saved in: {Style.RESET_ALL}{summary}\n")
                            else:
                                sys.stdout.write(f"{Fore.RED}Invalid summary file path.{Style.RESET_ALL}\n")
                        else:
                            sys.stdout.write(f"{Fore.RED}Invalid template folder path.{Style.RESET_ALL}\n")

                else:
                    sys.stdout.write(f"{Fore.RED}More than 2 sheets found in input sheet, Kindly provide 2 sheet in input.{Style.RESET_ALL}\n")

            else:
                sys.stdout.write(f"{Fore.RED}Invalid file path.{Style.RESET_ALL}\n")
        else:
            sys.stdout.write(f"{Fore.RED}Invalid file path, Kindly provide file with .xlsx extension.{Style.RESET_ALL}\n")

    except Exception as e:
        sys.stdout.write(f"{Fore.LIGHTRED_EX}Error Occurred: {Style.RESET_ALL}{e}\n")

    sys.stdout.write(f"{Fore.LIGHTMAGENTA_EX}{SIGN_OFF}{Style.RESET_ALL} \n")

    file1_path = input("Please provide the input file path: ").strip()
    file1_path = file1_path.replace('"', '')
    file1_path = str(PurePosixPath(PureWindowsPath(file1_path)))
